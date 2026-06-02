import io
from datetime import date, timedelta
from collections import defaultdict

from flask import Blueprint, send_file, g
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

from models.pedido import Pedido
from models.client import Client
from models.fiado import Fiado
from utils.auth_utils import require_auth, scoped_query

exportar_bp = Blueprint("exportar", __name__)

# ── Paleta NEXORA ──────────────────────────────────────────────────────────
_BLACK   = "FF0A0403"
_GOLD    = "FFF59E0B"
_GOLD2   = "FFFBBF24"
_CREAM   = "FFFEF3C7"
_ROW_ODD = "FF140800"
_ROW_EVN = "FF0A0403"
_BORDER  = "FF3D2000"

def _border(color=_BORDER):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _hdr_font():  return Font(bold=True, color=_GOLD, size=10)
def _sub_font():  return Font(bold=True, color=_CREAM, size=9)
def _body_font(): return Font(color=_CREAM, size=9)
def _hdr_fill():  return PatternFill("solid", fgColor=_BLACK)
def _odd_fill():  return PatternFill("solid", fgColor=_ROW_ODD)
def _evn_fill():  return PatternFill("solid", fgColor=_ROW_EVN)
def _gold_fill(): return PatternFill("solid", fgColor="FF1C0800")

def _center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def _left():   return Alignment(horizontal="left",   vertical="center", wrap_text=True)


@exportar_bp.route("/clientes/exportar-excel")
@require_auth
def exportar_excel():
    owner_id = g.owner_id

    # ── Datos ──────────────────────────────────────────────────────────────
    pedidos     = scoped_query(Pedido).order_by(Pedido.created_at.asc()).all()
    fiados      = scoped_query(Fiado).order_by(Fiado.created_at.asc()).all()
    clientes_db = scoped_query(Client).order_by(Client.nombre.asc()).all()

    # estructura: data[client_nombre][fecha_date] = {pedidos:[], fiados:[]}
    data = defaultdict(lambda: defaultdict(lambda: {"pedidos": [], "fiados": []}))
    todas_fechas = set()

    for p in pedidos:
        fecha_date = p.created_at.date()
        cant = int(p.cantidad) if p.cantidad == int(p.cantidad) else p.cantidad
        data[p.client_nombre][fecha_date]["pedidos"].append((p.product_nombre, cant, p.precio))
        todas_fechas.add(fecha_date)

    for f in fiados:
        fecha_date = f.created_at.date()
        data[f.client_nombre][fecha_date]["fiados"].append(
            (f.concepto, f.monto, f.pagado, f.pagado_at)
        )
        todas_fechas.add(fecha_date)

    fechas_ordenadas = sorted(todas_fechas)

    wb = Workbook()

    # ══════════════════════════════════════════════════════════════════════
    # HOJA 1 — CALENDARIO
    # ══════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Calendario"
    ws1.sheet_view.showGridLines = False
    ws1.freeze_panes = "B2"

    # Nombre de días en español
    DIAS = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"]

    # Fila 1 — encabezados de fecha
    ws1.cell(1, 1, "Cliente").font      = _hdr_font()
    ws1.cell(1, 1).fill      = _hdr_fill()
    ws1.cell(1, 1).alignment = _center()
    ws1.cell(1, 1).border    = _border()
    ws1.column_dimensions["A"].width = 24

    for col_idx, fecha in enumerate(fechas_ordenadas, start=2):
        dia_semana = DIAS[fecha.weekday()]
        label = f"{dia_semana}\n{fecha.strftime('%d/%m')}"
        c = ws1.cell(1, col_idx, label)
        c.font      = _hdr_font()
        c.fill      = _hdr_fill()
        c.alignment = _center()
        c.border    = _border()
        ws1.column_dimensions[get_column_letter(col_idx)].width = 18
        ws1.row_dimensions[1].height = 32

    # Filas de clientes
    clientes_con_datos = sorted(data.keys())
    for row_idx, nombre in enumerate(clientes_con_datos, start=2):
        fill = _odd_fill() if row_idx % 2 == 0 else _evn_fill()

        # Columna A — nombre
        c = ws1.cell(row_idx, 1, nombre)
        c.font      = Font(bold=True, color=_GOLD2, size=9)
        c.fill      = fill
        c.alignment = _left()
        c.border    = _border()

        total_row = 0
        for col_idx, fecha in enumerate(fechas_ordenadas, start=2):
            celda = data[nombre].get(fecha, None)
            if celda:
                lineas = []
                for prod, cant, precio in celda["pedidos"]:
                    lineas.append(f"{prod}  ×{cant}")
                    total_row += precio * cant
                for concepto, monto, pagado, pagado_at in celda["fiados"]:
                    estado = f"[PAGADO {pagado_at.strftime('%d/%m') if pagado_at else ''}]" if pagado else "[DEBE]"
                    lineas.append(f"💳 {concepto}  ${monto:,.0f} {estado}")
                texto = "\n".join(lineas)
            else:
                texto = ""

            c = ws1.cell(row_idx, col_idx, texto)
            c.font      = _body_font()
            c.fill      = fill
            c.alignment = _center()
            c.border    = _border()
            if texto:
                ws1.row_dimensions[row_idx].height = max(
                    ws1.row_dimensions[row_idx].height or 15,
                    15 * len(items)
                )

    # ══════════════════════════════════════════════════════════════════════
    # HOJA 2 — HISTORIAL DETALLADO
    # ══════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Historial")
    ws2.sheet_view.showGridLines = False
    ws2.freeze_panes = "A2"

    headers = ["Fecha", "Día", "Cliente", "Concepto / Producto", "Cantidad", "Monto / Precio", "Total", "Tipo", "Estado"]
    col_widths = [12, 8, 24, 28, 10, 14, 14, 10, 12]
    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        c = ws2.cell(1, col_idx, h)
        c.font      = _hdr_font()
        c.fill      = _hdr_fill()
        c.alignment = _center()
        c.border    = _border()
        ws2.column_dimensions[get_column_letter(col_idx)].width = w
    ws2.row_dimensions[1].height = 22

    # Combinar pedidos y fiados ordenados por fecha desc
    fiado_red  = Font(color="FFEF4444", size=9)
    fiado_green = Font(color="FF4ADE80", size=9)

    filas = []
    for p in pedidos:
        cant  = int(p.cantidad) if p.cantidad == int(p.cantidad) else p.cantidad
        total = p.precio * p.cantidad
        filas.append({
            "fecha":  p.created_at,
            "data":   [p.created_at.strftime("%d/%m/%Y"), DIAS[p.created_at.weekday()],
                       p.client_nombre, p.product_nombre, cant, p.precio, round(total,2), "Pedido", "—"],
            "tipo":   "pedido",
        })
    for f in fiados:
        pagado_str = f"Pagado {f.pagado_at.strftime('%d/%m/%Y')}" if f.pagado and f.pagado_at else ("Pagado" if f.pagado else "Pendiente")
        filas.append({
            "fecha":  f.created_at,
            "data":   [f.created_at.strftime("%d/%m/%Y"), DIAS[f.created_at.weekday()],
                       f.client_nombre, f.concepto, "—", f.monto, f.monto, "Fiado", pagado_str],
            "tipo":   "fiado",
            "pagado": f.pagado,
        })
    filas.sort(key=lambda x: x["fecha"], reverse=True)

    for row_idx, fila in enumerate(filas, start=2):
        fill = _odd_fill() if row_idx % 2 == 0 else _evn_fill()
        for col_idx, val in enumerate(fila["data"], start=1):
            c = ws2.cell(row_idx, col_idx, val)
            if fila["tipo"] == "fiado":
                c.font = fiado_green if fila.get("pagado") else fiado_red
            else:
                c.font = _body_font()
            c.fill      = fill
            c.alignment = _center() if col_idx not in (3, 4) else _left()
            c.border    = _border()
            if col_idx in (6, 7):
                c.number_format = '"$"#,##0.00'

    # ══════════════════════════════════════════════════════════════════════
    # HOJA 3 — RESUMEN POR CLIENTE
    # ══════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Resumen")
    ws3.sheet_view.showGridLines = False
    ws3.freeze_panes = "A2"

    h3 = ["Cliente", "Total pedidos", "Total facturado", "Último pedido"]
    w3 = [24, 14, 16, 14]
    for col_idx, (h, w) in enumerate(zip(h3, w3), start=1):
        c = ws3.cell(1, col_idx, h)
        c.font      = _hdr_font()
        c.fill      = _hdr_fill()
        c.alignment = _center()
        c.border    = _border()
        ws3.column_dimensions[get_column_letter(col_idx)].width = w
    ws3.row_dimensions[1].height = 22

    # Agrupar por cliente
    resumen = defaultdict(lambda: {"pedidos": 0, "total": 0.0, "ultimo": None})
    for p in pedidos:
        r = resumen[p.client_nombre]
        r["pedidos"] += 1
        r["total"]   += p.precio * p.cantidad
        if r["ultimo"] is None or p.created_at > r["ultimo"]:
            r["ultimo"] = p.created_at

    resumen_sorted = sorted(resumen.items(), key=lambda x: x[1]["total"], reverse=True)
    for row_idx, (nombre, r) in enumerate(resumen_sorted, start=2):
        fill = _odd_fill() if row_idx % 2 == 0 else _evn_fill()
        ultimo_str = r["ultimo"].strftime("%d/%m/%Y") if r["ultimo"] else "—"
        row_data = [nombre, r["pedidos"], round(r["total"], 2), ultimo_str]
        for col_idx, val in enumerate(row_data, start=1):
            c = ws3.cell(row_idx, col_idx, val)
            c.font      = _body_font()
            c.fill      = fill
            c.alignment = _left() if col_idx == 1 else _center()
            c.border    = _border()
            if col_idx == 3:
                c.number_format = '"$"#,##0.00'

    # ── Serializar y enviar ────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    hoy = date.today().strftime("%Y%m%d")
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"nexora_clientes_{hoy}.xlsx",
    )
